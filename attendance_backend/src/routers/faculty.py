from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from sqlalchemy import text, bindparam
import csv
import io
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
from src.core.database import engine
from src.core.utils import generate_code
from src.models.schemas import CreateClassRequest, StartSessionRequest, MarkAttendanceRequest, AdminResetPasswordRequest
from src import queries
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any
import os
import secrets
from src.core.config import RESET_ADMIN_KEY
from src.core.security import require_faculty, get_password_hash
from src.core.logging_config import get_logger

logger = get_logger("faculty")
router = APIRouter(tags=["faculty"])

# -------------------- FACULTY DASHBOARD --------------------

@router.get("/api/faculty/sessions/active")
async def get_active_sessions(faculty_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"⚡ [FACULTY/ACTIVE_SESSIONS] Fetch active sessions for faculty_id={faculty_id}")
    try:
        sql = text(
            """
            SELECT s.session_id, s.class_id, c.class_name, s.start_time, s.status
            FROM attendance_sessions s
            JOIN classes c ON s.class_id = c.class_id
            WHERE c.faculty_id = :faculty_id AND s.status = 'ACTIVE'
            ORDER BY s.start_time DESC
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"faculty_id": faculty_id})
            rows = [dict(r._mapping) for r in result]
            logger.info(f"✅ [FACULTY/ACTIVE_SESSIONS] Found {len(rows)} active sessions for faculty_id={faculty_id}")
            return rows
    except Exception as e:
        logger.exception(f"❌ [FACULTY/ACTIVE_SESSIONS] Error for faculty_id={faculty_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/faculty/{faculty_id}/classes")
async def get_faculty_classes(faculty_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"📚 [FACULTY/CLASSES] Fetching classes for faculty_id={faculty_id}")
    try:
        sql = text(
            """
            SELECT c.class_id, c.class_name, c.join_code,
                   COALESCE(ce.cnt, 0)::int AS students_count,
                   COALESCE(s.cnt, 0)::int  AS sessions_count,
                   s.last_session
            FROM classes c
            LEFT JOIN (
                SELECT class_id, COUNT(*) AS cnt
                FROM class_enrollments
                GROUP BY class_id
            ) ce ON ce.class_id = c.class_id
            LEFT JOIN (
                SELECT class_id, COUNT(*) AS cnt, MAX(start_time) AS last_session
                FROM attendance_sessions
                GROUP BY class_id
            ) s ON s.class_id = c.class_id
            WHERE c.faculty_id = :faculty_id
            ORDER BY c.class_name
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"faculty_id": faculty_id})
            rows = [dict(r._mapping) for r in result]
            logger.info(f"✅ [FACULTY/CLASSES] Found {len(rows)} classes for faculty_id={faculty_id}")
            return rows
    except Exception as e:
        logger.exception(f"❌ [FACULTY/CLASSES] Error fetching classes for faculty_id={faculty_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/faculty/classes")
async def create_faculty_class(class_data: CreateClassRequest, current_user: dict = Depends(require_faculty)):
    logger.info(f"➕ [FACULTY/CREATE_CLASS] Creating class '{class_data.class_name}' for faculty_id={class_data.faculty_id}")
    try:
        async with engine.begin() as conn:
            # Check if a class with the same name already exists for this faculty
            check_sql = text(
                """
                SELECT class_id FROM classes 
                WHERE class_name = :class_name AND faculty_id = :faculty_id
                LIMIT 1
                """
            )
            existing = await conn.execute(
                check_sql,
                {
                    "class_name": class_data.class_name,
                    "faculty_id": class_data.faculty_id,
                }
            )
            if existing.fetchone():
                logger.warning(f"⚠️ [FACULTY/CREATE_CLASS] Duplicate class name '{class_data.class_name}' for faculty_id={class_data.faculty_id}")
                raise HTTPException(
                    status_code=400,
                    detail=f"A class with the name '{class_data.class_name}' already exists"
                )
            
            join_code = class_data.join_code or generate_code()
            sql = text(
                """
                INSERT INTO classes (class_name, faculty_id, join_code)
                VALUES (:class_name, :faculty_id, :join_code)
                RETURNING class_id, class_name, join_code
                """
            )
            res = await conn.execute(
                sql,
                {
                    "class_name": class_data.class_name,
                    "faculty_id": class_data.faculty_id,
                    "join_code": join_code,
                },
            )
            created_class = dict(res.fetchone()._mapping)
            logger.info(f"✅ [FACULTY/CREATE_CLASS] Created class_id={created_class['class_id']} ('{created_class['class_name']}')")
            return created_class
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [FACULTY/CREATE_CLASS] Failed to create class '{class_data.class_name}': {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/api/faculty/classes/{class_id}")
async def delete_faculty_class(class_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"🗑️ [FACULTY/DELETE_CLASS] Deleting class_id={class_id}")
    try:
        sql = text("DELETE FROM classes WHERE class_id = :class_id RETURNING class_id")
        async with engine.begin() as conn:
            res = await conn.execute(sql, {"class_id": class_id})
            if res.rowcount == 0:
                logger.warning(f"⚠️ [FACULTY/DELETE_CLASS] Class not found: class_id={class_id}")
                raise HTTPException(status_code=404, detail="Class not found")
            logger.info(f"✅ [FACULTY/DELETE_CLASS] Successfully deleted class_id={class_id}")
            return {"message": "Class deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [FACULTY/DELETE_CLASS] Failed to delete class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/faculty/classes/{class_id}/sessions")
async def start_session(class_id: int, request: StartSessionRequest = None, current_user: dict = Depends(require_faculty)):
    """Start a new attendance session with generated code and optional location"""
    logger.info(f"🚀 [FACULTY/START_SESSION] Starting attendance session for class_id={class_id}")
    try:
        if request is None:
            request = StartSessionRequest(class_id=class_id)
        
        async with engine.begin() as conn:
            # Check for existing active session
            check_sql = text(
                """
                SELECT session_id, generated_code 
                FROM attendance_sessions 
                WHERE class_id = :class_id AND STATUS = 'ACTIVE'
                LIMIT 1
                """
            )
            result = await conn.execute(check_sql, {"class_id": class_id})
            existing = result.fetchone()
            
            if existing:
                existing_data = dict(existing._mapping)
                logger.info(f"ℹ️ [FACULTY/START_SESSION] Active session already exists: session_id={existing_data['session_id']}")
                return existing_data
            
            code = generate_code()
            
            # UTC+5:30
            utc_now = datetime.utcnow()
            ist_offset = timedelta(hours=5, minutes=30)
            current_time_ist = utc_now + ist_offset
            
            # Insert with location columns
            sql = text(
                """
                INSERT INTO attendance_sessions (class_id, start_time, status, generated_code, latitude, longitude, radius_meters)
                VALUES (:class_id, :start_time, 'ACTIVE', :code, :lat, :lon, :rad)
                RETURNING session_id, class_id, start_time, status, generated_code, latitude, longitude, radius_meters
                """
            )
            res = await conn.execute(sql, {
                "class_id": class_id, 
                "start_time": current_time_ist, 
                "code": code,
                "lat": request.latitude,
                "lon": request.longitude,
                "rad": request.radius_meters
            })
            
            session_data = dict(res.fetchone()._mapping)
            logger.info(f"✅ [FACULTY/START_SESSION] Session started: session_id={session_data['session_id']}, code={code}")
            return session_data
    except Exception as e:
        logger.exception(f"❌ [FACULTY/START_SESSION] Error starting session for class_id={class_id}: {e}")
        if "column" in str(e).lower() and ("latitude" in str(e).lower() or "longitude" in str(e).lower()):
             raise HTTPException(status_code=500, detail="Database schema outdated. Please run migration to add location columns.")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/api/faculty/classes/{class_id}/sessions/{session_id}/end")
async def end_session(class_id: int, session_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"🛑 [FACULTY/END_SESSION] Ending session_id={session_id} for class_id={class_id}")
    try:
        async with engine.begin() as conn:
            # Mark absent students
            mark_absent_sql = text(
                """
                INSERT INTO attendance_records (session_id, student_id, status, marked_at)
                SELECT :session_id, ce.student_id, 'ABSENT', NOW()
                FROM class_enrollments ce
                WHERE ce.class_id = :class_id
                AND NOT EXISTS (
                    SELECT 1 FROM attendance_records ar
                    WHERE ar.session_id = :session_id
                    AND ar.student_id = ce.student_id
                )
                """
            )
            await conn.execute(mark_absent_sql, {"session_id": session_id, "class_id": class_id})
            
            # Update session status
            utc_now = datetime.utcnow()
            ist_offset = timedelta(hours=5, minutes=30)
            current_time_ist = utc_now + ist_offset
            
            sql = text(
                """
                UPDATE attendance_sessions
                SET end_time = :end_time, status = 'CLOSED'
                WHERE session_id = :session_id AND class_id = :class_id
                RETURNING *
                """
            )
            result = await conn.execute(sql, {"session_id": session_id, "class_id": class_id, "end_time": current_time_ist})
            row = result.fetchone()
            
            if not row:
                logger.warning(f"⚠️ [FACULTY/END_SESSION] Session not found: session_id={session_id}")
                raise HTTPException(status_code=404, detail="Session not found")
            
            logger.info(f"✅ [FACULTY/END_SESSION] Closed session_id={session_id} for class_id={class_id}")
            return dict(row._mapping)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [FACULTY/END_SESSION] Error closing session_id={session_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ... Additional endpoints ...

@router.get("/api/faculty/classes/{class_id}/sessions/by-date")
async def get_sessions_by_date_endpoint(class_id: int, date: str, current_user: dict = Depends(require_faculty)):
    logger.info(f"📅 [FACULTY/SESSIONS_BY_DATE] Class ID={class_id}, Date={date}")
    try:
        sql = text(
            """
            SELECT session_id, start_time, end_time, status, generated_code
            FROM attendance_sessions
            WHERE class_id = :class_id 
              AND TO_CHAR(start_time, 'YYYY-MM-DD') = :date
            ORDER BY start_time ASC
            """
        )
        
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"class_id": class_id, "date": date})
            return [dict(r._mapping) for r in result]
    except Exception as e:
        logger.exception(f"❌ [FACULTY/SESSIONS_BY_DATE] Error for class_id={class_id}, date={date}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/faculty/classes/{class_id}/session-dates")
async def get_class_session_dates(class_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"📅 [FACULTY/SESSION_DATES] Fetch session dates for class_id={class_id}")
    try:
        sql = text(
            """
            SELECT 
                TO_CHAR(start_time, 'YYYY-MM-DD') as date,
                COUNT(session_id) as session_count,
                MAX(start_time) as latest_start_time
            FROM attendance_sessions
            WHERE class_id = :class_id
            GROUP BY TO_CHAR(start_time, 'YYYY-MM-DD')
            ORDER BY date DESC
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"class_id": class_id})
            return [dict(r._mapping) for r in result]
    except Exception as e:
        logger.exception(f"❌ [FACULTY/SESSION_DATES] Error for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/faculty/classes/{class_id}/sessions/stats")
async def get_class_sessions_stats(class_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"📊 [FACULTY/SESSION_STATS] Fetch session stats for class_id={class_id}")
    try:
        sql = text(
            """
            SELECT
                COUNT(*)::int AS sessions_count,
                MAX(start_time) AS last_session
            FROM attendance_sessions
            WHERE class_id = :class_id
            """
        )

        async with engine.connect() as conn:
            result = await conn.execute(sql, {"class_id": class_id})
            row = result.fetchone()
            if not row:
                return {"sessions_count": 0, "last_session": None}
            return dict(row._mapping)
    except Exception as e:
        logger.exception(f"❌ [FACULTY/SESSION_STATS] Error for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions/{date}")
async def sessions_by_date(date: str, current_user: dict = Depends(require_faculty)):
    logger.info(f"📅 [FACULTY/SESSIONS_BY_DATE_LEGACY] Date={date}")
    try:
        return await queries.get_sessions_by_date(date)
    except Exception as e:
        logger.exception(f"❌ [FACULTY/SESSIONS_BY_DATE_LEGACY] Error for date={date}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/session/{session_id}/attendance")
async def attendance_for_session(session_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"📋 [FACULTY/ATTENDANCE_SESSION] Session ID={session_id}")
    try:
        return await queries.get_attendance_for_session(session_id)
    except Exception as e:
        logger.exception(f"❌ [FACULTY/ATTENDANCE_SESSION] Error for session_id={session_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/class/{class_id}/absent/{date}")
async def absent_students(class_id: int, date: str, current_user: dict = Depends(require_faculty)):
    logger.info(f"❌ [FACULTY/ABSENT_STUDENTS] Class ID={class_id}, Date={date}")
    try:
        return await queries.get_absent_students_in_class_on_date(class_id, date)
    except Exception as e:
        logger.exception(f"❌ [FACULTY/ABSENT_STUDENTS] Error for class_id={class_id}, date={date}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/class/{class_id}/students/below_percentage")
async def students_below_percentage(class_id: int, threshold: Optional[float] = 75.0, current_user: dict = Depends(require_faculty)):
    logger.info(f"⚠️ [FACULTY/STUDENTS_BELOW_PCT] Class ID={class_id}, Threshold={threshold}")
    try:
        return await queries.get_students_below_percentage(class_id, threshold)
    except Exception as e:
        logger.exception(f"❌ [FACULTY/STUDENTS_BELOW_PCT] Error for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/most-active-class")
async def most_active_class(current_user: dict = Depends(require_faculty)):
    logger.info("🏆 [FACULTY/MOST_ACTIVE_CLASS] Requesting most active class")
    try:
        result = await queries.get_most_active_class()
        if not result:
            logger.warning("⚠️ [FACULTY/MOST_ACTIVE_CLASS] No classes or attendance records found")
            raise HTTPException(status_code=404, detail="No classes or attendance records found")
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [FACULTY/MOST_ACTIVE_CLASS] Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/faculty-with-classes")
async def faculty_with_classes(current_user: dict = Depends(require_faculty)):
    logger.info("👥 [FACULTY/WITH_CLASSES] Listing faculty with classes")
    try:
        return await queries.get_faculty_with_classes()
    except Exception as e:
        logger.exception(f"❌ [FACULTY/WITH_CLASSES] Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/faculty/classes/{class_id}/attendance")
async def get_session_attendance(class_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"📋 [FACULTY/CLASS_ATTENDANCE] Fetch attendance for class_id={class_id}")
    try:
        sql = text(
            """
            SELECT
                s.session_id, s.start_time, s.end_time, s.status,
                u.user_id as student_id,
                u.name AS student_name,
                ar.status AS attendance_status,
                ar.marked_at
            FROM attendance_sessions s
            LEFT JOIN attendance_records ar ON s.session_id = ar.session_id
            LEFT JOIN users u ON ar.student_id = u.user_id
            WHERE s.class_id = :class_id
            ORDER BY s.start_time DESC, u.name
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"class_id": class_id})
            return [dict(r._mapping) for r in result]
    except Exception as e:
        logger.exception(f"❌ [FACULTY/CLASS_ATTENDANCE] Error for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/class/{class_id}/active-session")
async def get_active_session(class_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"🔍 [FACULTY/ACTIVE_SESSION_CHECK] Class ID={class_id}")
    try:
        sql = text(
            """
            SELECT session_id, class_id, start_time, status, generated_code
            FROM attendance_sessions
            WHERE class_id = :class_id AND status = 'ACTIVE'
            ORDER BY start_time DESC
            LIMIT 1
            """
        )
        async with engine.connect() as conn:
            row = await conn.execute(sql, {"class_id": class_id})
            res = row.fetchone()
            return dict(res._mapping) if res else {}
    except Exception as e:
        logger.exception(f"❌ [FACULTY/ACTIVE_SESSION_CHECK] Error for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/faculty/sessions/{session_id}")
async def get_session_by_id(session_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"🔍 [FACULTY/GET_SESSION] Fetch session_id={session_id}")
    try:
        sql = text("SELECT * FROM attendance_sessions WHERE session_id = :session_id")
        async with engine.connect() as conn:
            row = await conn.execute(sql, {"session_id": session_id})
            res = row.fetchone()
            if not res:
                logger.warning(f"⚠️ [FACULTY/GET_SESSION] Session not found: session_id={session_id}")
                raise HTTPException(status_code=404, detail="Session not found")
            return dict(res._mapping)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [FACULTY/GET_SESSION] Error for session_id={session_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/api/faculty/sessions/{session_id}")
async def delete_session(session_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"🗑️ [FACULTY/DELETE_SESSION] Deleting session_id={session_id}")
    try:
        async with engine.begin() as conn:
            owner_sql = text(
                """
                SELECT s.session_id, c.faculty_id
                FROM attendance_sessions s
                JOIN classes c ON s.class_id = c.class_id
                WHERE s.session_id = :session_id
                """
            )
            sess_row = (await conn.execute(owner_sql, {"session_id": session_id})).fetchone()
            if not sess_row:
                logger.warning(f"⚠️ [FACULTY/DELETE_SESSION] Session not found: session_id={session_id}")
                raise HTTPException(status_code=404, detail="Session not found")

            if sess_row.faculty_id != current_user["user_id"]:
                logger.warning(f"⚠️ [FACULTY/DELETE_SESSION] Access denied for user_id={current_user['user_id']}")
                raise HTTPException(status_code=403, detail="Access denied")

            del_records_sql = text("DELETE FROM attendance_records WHERE session_id = :session_id")
            await conn.execute(del_records_sql, {"session_id": session_id})

            del_session_sql = text("DELETE FROM attendance_sessions WHERE session_id = :session_id")
            await conn.execute(del_session_sql, {"session_id": session_id})

            logger.info(f"✅ [FACULTY/DELETE_SESSION] Deleted session_id={session_id} and attendance records")
            return {"message": "Session and associated attendance records deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [FACULTY/DELETE_SESSION] Error deleting session_id={session_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/api/faculty/classes/{class_id}/students")
async def get_class_students(class_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"👥 [FACULTY/CLASS_STUDENTS] Fetch students for class_id={class_id}")
    try:
        sql = text(
            """
            SELECT u.user_id, u.name, u.email, ce.roll_number, ce.section
            FROM class_enrollments ce
            JOIN users u ON ce.student_id = u.user_id
            WHERE ce.class_id = :class_id
            ORDER BY ce.roll_number, u.name
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"class_id": class_id})
            rows = [dict(r._mapping) for r in result]
            logger.info(f"✅ [FACULTY/CLASS_STUDENTS] Found {len(rows)} enrolled students for class_id={class_id}")
            return rows
    except Exception as e:
        logger.exception(f"❌ [FACULTY/CLASS_STUDENTS] Error for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/faculty/classes/{class_id}/details")
async def faculty_class_details(class_id: int, current_user: dict = Depends(require_faculty)):
    logger.info(f"🔍 [FACULTY/CLASS_DETAILS] Fetch class details for class_id={class_id}")
    try:
        sql = text(
            """
            SELECT c.class_id, c.class_name, c.join_code, u.name AS faculty_name
            FROM classes c
            JOIN users u ON c.faculty_id = u.user_id
            WHERE c.class_id = :cid
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"cid": class_id})
            row = result.fetchone()
            if not row:
                logger.warning(f"⚠️ [FACULTY/CLASS_DETAILS] Class not found: class_id={class_id}")
                raise HTTPException(status_code=404, detail="Class not found")
            return dict(row._mapping)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [FACULTY/CLASS_DETAILS] Error for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/session/{session_id}/attendance")
async def mark_attendance_manual(session_id: int, payload: MarkAttendanceRequest, current_user: dict = Depends(require_faculty)):
    logger.info(f"✍️ [FACULTY/MARK_ATTENDANCE_MANUAL] Manually marking student_id={payload.student_id} as '{payload.status}' in session_id={session_id}")
    try:
        status = (payload.status or "PRESENT").upper()
        if status not in ("PRESENT", "LATE", "ABSENT"):
            logger.warning(f"⚠️ [FACULTY/MARK_ATTENDANCE_MANUAL] Invalid status '{status}' requested")
            raise HTTPException(status_code=400, detail="Invalid status")

        async with engine.begin() as conn:
            # Check session
            s = (await conn.execute(text("SELECT 1 FROM attendance_sessions WHERE session_id = :sid"), {"sid": session_id})).fetchone()
            if not s:
                 logger.warning(f"⚠️ [FACULTY/MARK_ATTENDANCE_MANUAL] Session not found: session_id={session_id}")
                 raise HTTPException(status_code=404, detail="Session not found")
            
            # Upsert
            upd = await conn.execute(
                text("UPDATE attendance_records SET status = :st, marked_at = NOW() WHERE session_id = :sid AND student_id = :uid"),
                {"st": status, "sid": session_id, "uid": payload.student_id}
            )
            if upd.rowcount == 0:
                 await conn.execute(
                    text("INSERT INTO attendance_records (session_id, student_id, status, marked_at) VALUES (:sid, :uid, :st, NOW())"),
                    {"sid": session_id, "uid": payload.student_id, "st": status}
                 )
            
            logger.info(f"✅ [FACULTY/MARK_ATTENDANCE_MANUAL] Updated student_id={payload.student_id} to status='{status}' for session_id={session_id}")
            return {"message": "Attendance updated", "status": status}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [FACULTY/MARK_ATTENDANCE_MANUAL] Error for session_id={session_id}, student_id={payload.student_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/faculty/sessions/{session_id}/attendance/flat")
async def get_session_attendance_flat(session_id: int, current_user: dict = Depends(require_faculty)):
    """
    Get attendance for a specific session in a flat format suitable for tables.
    Includes all enrolled students and their status (PRESENT/ABSENT/LATE) for this session.
    """
    logger.info(f"📜 [FACULTY/SESSION_ATTENDANCE_FLAT] Session ID={session_id}")
    try:
        sql = text(
            """
            SELECT 
                ce.student_id,
                u.name as student_name,
                ce.roll_number,
                ce.section,
                COALESCE(ar.status, 'ABSENT') as status,
                ar.marked_at
            FROM class_enrollments ce
            JOIN attendance_sessions s ON s.class_id = ce.class_id
            JOIN users u ON ce.student_id = u.user_id
            LEFT JOIN attendance_records ar ON ar.session_id = s.session_id AND ar.student_id = ce.student_id
            WHERE s.session_id = :session_id
            ORDER BY ce.roll_number
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"session_id": session_id})
            return [dict(r._mapping) for r in result]
    except Exception as e:
        logger.exception(f"❌ [FACULTY/SESSION_ATTENDANCE_FLAT] Error for session_id={session_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/faculty/classes/{class_id}/students/attendance-stats")
async def get_students_attendance_stats(
    class_id: int,
    min_pct: Optional[float] = None,
    max_pct: Optional[float] = None,
    current_user: dict = Depends(require_faculty)
):
    """
    Returns all enrolled students in a class with their overall attendance percentage.
    Optionally filter by min_pct and/or max_pct (0-100).
    """
    logger.info(f"📊 [FACULTY/STUDENTS_ATTENDANCE_STATS] Class ID={class_id}, min_pct={min_pct}, max_pct={max_pct}")
    try:
        sql = text(
            """
            SELECT
                u.user_id AS student_id,
                u.name AS student_name,
                u.email,
                ce.roll_number,
                ce.section,
                COUNT(s.session_id) AS total_sessions,
                COUNT(CASE WHEN ar.status IN ('PRESENT', 'LATE') THEN 1 END) AS present_count,
                CASE
                    WHEN COUNT(s.session_id) = 0 THEN 0.0
                    ELSE ROUND(
                        COUNT(CASE WHEN ar.status IN ('PRESENT', 'LATE') THEN 1 END) * 100.0
                        / COUNT(s.session_id),
                        2
                    )
                END AS attendance_percentage
            FROM class_enrollments ce
            JOIN users u ON ce.student_id = u.user_id
            CROSS JOIN (
                SELECT session_id FROM attendance_sessions WHERE class_id = :class_id
            ) s
            LEFT JOIN attendance_records ar
                ON ar.session_id = s.session_id AND ar.student_id = ce.student_id
            WHERE ce.class_id = :class_id
            GROUP BY u.user_id, u.name, u.email, ce.roll_number, ce.section
            ORDER BY ce.section, ce.roll_number, u.name
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql, {"class_id": class_id})
            rows = [dict(r._mapping) for r in result]

        if min_pct is not None:
            rows = [r for r in rows if r["attendance_percentage"] >= min_pct]
        if max_pct is not None:
            rows = [r for r in rows if r["attendance_percentage"] <= max_pct]

        return rows
    except Exception as e:
        logger.exception(f"❌ [FACULTY/STUDENTS_ATTENDANCE_STATS] Error for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- FACULTY ADMIN: RESET PASSWORD --------------------

@router.get("/api/faculty/users")
async def list_all_users(current_user: dict = Depends(require_faculty)):
    """List all users (students + faculty) for the password reset picker."""
    logger.info("👥 [FACULTY/LIST_USERS] Listing all users for admin reset picker")
    try:
        sql = text(
            """
            SELECT user_id, name, email, role
            FROM users
            ORDER BY role, name
            """
        )
        async with engine.connect() as conn:
            result = await conn.execute(sql)
            return [dict(r._mapping) for r in result]
    except Exception as e:
        logger.exception(f"❌ [FACULTY/LIST_USERS] Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/faculty/admin/reset-password")
async def admin_reset_password(request: AdminResetPasswordRequest, current_user: dict = Depends(require_faculty)):
    """Faculty-only: directly reset any user's password (no email token required)."""
    logger.info(f"🔑 [ADMIN_RESET] Admin password reset requested for target user_id={request.user_id}")
    try:
        server_admin_key = RESET_ADMIN_KEY
        if not server_admin_key:
            logger.error("❌ [ADMIN_RESET] SECURITY ALERT: RESET_ADMIN_KEY is not configured in the application config!")
            raise HTTPException(
                status_code=500,
                detail="Password reset is disabled: Admin reset key is not configured."
            )
        
        if not secrets.compare_digest(request.admin_key, server_admin_key):
            logger.warning(f"⚠️ [ADMIN_RESET] Invalid admin reset key attempt by user_id={current_user.get('user_id')}")
            raise HTTPException(
                status_code=403,
                detail="Invalid admin reset key. Access denied."
            )

        if len(request.new_password) < 6:
            raise HTTPException(
                status_code=400,
                detail="Password must be at least 6 characters long"
            )

        from src.core.security import get_password_hash
        new_hash = get_password_hash(request.new_password)

        async with engine.begin() as conn:
            check_sql = text("SELECT user_id, name, email FROM users WHERE user_id = :user_id")
            result = await conn.execute(check_sql, {"user_id": request.user_id})
            user = result.fetchone()
            if not user:
                logger.warning(f"⚠️ [ADMIN_RESET] User not found: user_id={request.user_id}")
                raise HTTPException(status_code=404, detail="User not found")

            update_sql = text(
                "UPDATE users SET password_hash = :password_hash WHERE user_id = :user_id"
            )
            await conn.execute(update_sql, {
                "password_hash": new_hash,
                "user_id": request.user_id
            })

        user_data = dict(user._mapping)
        logger.info(f"✅ [ADMIN_RESET] Password reset successfully for user_id={request.user_id} ({user_data['email']})")
        return {
            "message": f"Password for {user_data['name']} has been reset successfully",
            "success": True
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ [ADMIN_RESET] Error resetting password for user_id={request.user_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/faculty/classes/{class_id}/sessions/all-with-attendance")
async def get_all_sessions_with_attendance(class_id: int, current_user: dict = Depends(require_faculty)):
    """
    Get all sessions with their attendance records flat, optimized for a single export file.
    """
    logger.info(f"📦 [FACULTY/EXPORT_ALL_SESSIONS] Exporting all sessions for class_id={class_id}")
    try:
        sessions_sql = text(
            """
            SELECT session_id, start_time, end_time, status, generated_code
            FROM attendance_sessions
            WHERE class_id = :class_id
            ORDER BY start_time DESC
            """
        )
        
        records_sql = text(
            """
            SELECT 
                s.session_id,
                ce.student_id,
                u.name as student_name,
                u.email as student_email,
                ce.roll_number,
                ce.section,
                COALESCE(ar.status, 'ABSENT') as status,
                ar.marked_at
            FROM attendance_sessions s
            JOIN class_enrollments ce ON s.class_id = ce.class_id
            JOIN users u ON ce.student_id = u.user_id
            LEFT JOIN attendance_records ar ON ar.session_id = s.session_id AND ar.student_id = ce.student_id
            WHERE s.class_id = :class_id
            ORDER BY ce.roll_number
            """
        )
        
        async with engine.connect() as conn:
            sessions_res = await conn.execute(sessions_sql, {"class_id": class_id})
            sessions_rows = [dict(r._mapping) for r in sessions_res]
            
            records_res = await conn.execute(records_sql, {"class_id": class_id})
            records_rows = [dict(r._mapping) for r in records_res]
            
        records_by_session = {}
        for r in records_rows:
            sid = r["session_id"]
            if sid not in records_by_session:
                records_by_session[sid] = []
            
            marked_at_str = r["marked_at"].isoformat() if r["marked_at"] else None
            
            records_by_session[sid].append({
                "student_id": r["student_id"],
                "student_name": r["student_name"],
                "email": r["student_email"],
                "roll_number": r["roll_number"],
                "section": r["section"],
                "status": r["status"],
                "marked_at": marked_at_str
            })
            
        sessions_list = []
        for s in sessions_rows:
            sid = s["session_id"]
            recs = records_by_session.get(sid, [])
            
            present_count = sum(1 for r in recs if r["status"] == "PRESENT")
            late_count = sum(1 for r in recs if r["status"] == "LATE")
            absent_count = sum(1 for r in recs if r["status"] == "ABSENT")
            
            sessions_list.append({
                "session_id": sid,
                "start_time": s["start_time"].isoformat() if s["start_time"] else None,
                "end_time": s["end_time"].isoformat() if s["end_time"] else None,
                "status": s["status"],
                "generated_code": s["generated_code"],
                "records": recs,
                "totals": {
                    "present": present_count,
                    "late": late_count,
                    "absent": absent_count
                }
            })
            
        logger.info(f"✅ [FACULTY/EXPORT_ALL_SESSIONS] Successfully processed {len(sessions_list)} sessions for class_id={class_id}")
        return {"sessions": sessions_list}
    except Exception as e:
        logger.exception(f"❌ [FACULTY/EXPORT_ALL_SESSIONS] Error exporting sessions for class_id={class_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/faculty/classes/{class_id}/bulk-register-students")
async def bulk_register_students(
    class_id: int,
    file: UploadFile = File(...),
    default_password: Optional[str] = Form("Student@123"),
    current_user: dict = Depends(require_faculty)
):
    """Bulk register students from CSV or Excel file and enroll them into a class with default password and mandatory first-login password change."""
    logger.info(f"👥 [FACULTY/BULK_REGISTER] Bulk registering students for class_id={class_id}, filename={file.filename}")
    
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    filename = file.filename.lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Unsupported file format. Please upload a .csv or .xlsx file.")
    
    content = await file.read()
    raw_rows = []
    
    try:
        if filename.endswith(".csv"):
            text_content = content.decode("utf-8-sig", errors="ignore")
            csv_reader = csv.DictReader(io.StringIO(text_content))
            for row in csv_reader:
                raw_rows.append({str(k).strip().lower(): str(v).strip() for k, v in row.items() if k is not None})
        else:
            if not HAS_OPENPYXL:
                raise HTTPException(status_code=400, detail="The openpyxl library is required on backend to process .xlsx files. Please upload a .csv file or install openpyxl.")
            wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
            sheet = wb.active
            iter_rows = list(sheet.iter_rows(values_only=True))
            if iter_rows:
                headers = [str(h).strip().lower() if h is not None else "" for h in iter_rows[0]]
                for r in iter_rows[1:]:
                    if not r or not any(r):
                        continue
                    row_dict = {}
                    for idx, val in enumerate(r):
                        if idx < len(headers) and headers[idx]:
                            row_dict[headers[idx]] = str(val).strip() if val is not None else ""
                    raw_rows.append(row_dict)
    except Exception as e:
        logger.exception(f"❌ [FACULTY/BULK_REGISTER] Failed to parse file {file.filename}: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse spreadsheet file: {str(e)}")
    
    if not raw_rows:
        raise HTTPException(status_code=400, detail="The uploaded file contains no data rows.")
    
    def get_field(row: dict, keys: list) -> str:
        for k in keys:
            if k in row and row[k]:
                return row[k].strip()
        return ""

    created_users_count = 0
    enrolled_count = 0
    updated_enrollment_count = 0
    errors = []
    
    default_pwd = default_password or "Student@123"
    hashed_default_pwd = get_password_hash(default_pwd)
    
    async with engine.begin() as conn:
        # Verify class exists
        class_check = await conn.execute(
            text("SELECT class_id FROM classes WHERE class_id = :class_id"),
            {"class_id": class_id}
        )
        if not class_check.fetchone():
            raise HTTPException(status_code=404, detail="Class not found")
        
        for idx, row in enumerate(raw_rows, start=2):  # row 1 is header
            name = get_field(row, ["name", "student_name", "student name", "full_name", "full name"])
            email = get_field(row, ["email", "student_email", "student email", "mail"]).lower()
            roll_number = get_field(row, ["roll_number", "roll number", "roll_no", "roll no", "rollno", "id"])
            section = get_field(row, ["section", "sec"])
            
            if not email or not name:
                errors.append(f"Row {idx}: Missing email or name (name='{name}', email='{email}')")
                continue
            
            # 1. Check or Create User
            user_res = await conn.execute(
                text("SELECT user_id FROM users WHERE LOWER(email) = :email"),
                {"email": email}
            )
            user_row = user_res.fetchone()
            
            if user_row:
                user_id = user_row[0]
            else:
                # Create user with default password and must_change_password = TRUE
                ins_res = await conn.execute(
                    text("""
                        INSERT INTO users (name, email, password_hash, role, must_change_password)
                        VALUES (:name, :email, :password_hash, 'STUDENT', TRUE)
                        RETURNING user_id
                    """),
                    {
                        "name": name,
                        "email": email,
                        "password_hash": hashed_default_pwd
                    }
                )
                user_id = ins_res.fetchone()[0]
                created_users_count += 1
            
            # 2. Check or Enroll in Class
            enroll_res = await conn.execute(
                text("SELECT enrollment_id FROM class_enrollments WHERE class_id = :class_id AND student_id = :student_id"),
                {"class_id": class_id, "student_id": user_id}
            )
            enroll_row = enroll_res.fetchone()
            
            if enroll_row:
                # Update existing enrollment details
                await conn.execute(
                    text("""
                        UPDATE class_enrollments 
                        SET roll_number = COALESCE(NULLIF(:roll_number, ''), roll_number),
                            section = COALESCE(NULLIF(:section, ''), section)
                        WHERE class_id = :class_id AND student_id = :student_id
                    """),
                    {
                        "class_id": class_id,
                        "student_id": user_id,
                        "roll_number": roll_number,
                        "section": section
                    }
                )
                updated_enrollment_count += 1
            else:
                # Create new enrollment
                await conn.execute(
                    text("""
                        INSERT INTO class_enrollments (class_id, student_id, roll_number, section)
                        VALUES (:class_id, :student_id, :roll_number, :section)
                    """),
                    {
                        "class_id": class_id,
                        "student_id": user_id,
                        "roll_number": roll_number,
                        "section": section
                    }
                )
                enrolled_count += 1

    logger.info(f"✅ [FACULTY/BULK_REGISTER] Bulk register complete for class_id={class_id}: Created {created_users_count} users, enrolled {enrolled_count} new, updated {updated_enrollment_count}")
    return {
        "message": f"Successfully processed {len(raw_rows)} rows.",
        "total_rows": len(raw_rows),
        "created_users_count": created_users_count,
        "enrolled_count": enrolled_count,
        "updated_enrollment_count": updated_enrollment_count,
        "default_password": default_pwd,
        "errors": errors
    }



